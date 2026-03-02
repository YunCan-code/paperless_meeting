from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func
from typing import List, Optional
import io
import openpyxl
from urllib.parse import quote
from database import get_session
from models import User, UserRead
from utils.security import hash_password, verify_password

# Create Router
router = APIRouter(prefix="/users", tags=["users"])

# Response Model using Pydantic Generic not supported directly in simple setup, so we return dict or specific schema
# For list with pagination, we can return { "items": [...], "total": ... }

from pydantic import BaseModel

class ChangePasswordRequest(BaseModel):
    user_id: int
    old_password: str
    new_password: str

@router.get("/stats")
def get_user_stats(session: Session = Depends(get_session)):
    """
    Get User Statistics
    """
    total = session.exec(select(func.count(User.id))).one()
    active_users = session.exec(select(func.count(User.id)).where(User.is_active == True)).one()
    
    # Calculate departments count (distinct)
    departments = session.exec(select(func.count(func.distinct(User.department))).where(User.department != None, User.department != "")).one()
    
    # Calculate active_today
    # Logic: last_login >= today 00:00
    from datetime import datetime
    now = datetime.now()
    today_start = datetime(now.year, now.month, now.day)
    
    active_today = session.exec(select(func.count(User.id)).where(User.last_login >= today_start)).one()
    
    return {
        "total": total,
        "active_users": active_users,
        "departments": departments,
        "active_today": active_today
    }

@router.get("/template")
def download_template():
    """
    Download User Import Template
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    # Headers
    headers = ["姓名(*必填)", "区县", "部门", "手机号(*必填)", "密码(*必填)", "邮箱"]
    ws.append(headers)
    
    # Sample Row
    ws.append(["张三", "五华区", "研发部", "13800138000", "123456", "zhangsan@example.com"])
    
    # Adjust column width
    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = quote("用户导入模板.xlsx")
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}; filename*=utf-8''{filename}"}
    )

@router.post("/import")
async def import_users(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Import Users from Excel
    """
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        users_to_add = []
        errors = []
        seen_phones = {}  # 文件内手机号去重: phone -> 行号

        # Skip header, iterate rows
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Row: Name, District, Dept, Phone, Password, Email
            padded = list(row) + [None] * max(0, 6 - len(row))
            name, district, dept, phone, password, email = padded[0:6]

            # 校验必填项
            has_error = False
            if not name:
                errors.append(f"第 {i} 行: 姓名不能为空")
                has_error = True
            if not phone:
                errors.append(f"第 {i} 行: 手机号不能为空")
                has_error = True
            if not password:
                errors.append(f"第 {i} 行: 密码不能为空")
                has_error = True

            if has_error:
                continue

            # 手机号查重
            phone_str = str(phone).strip()

            # 文件内互查
            if phone_str in seen_phones:
                errors.append(f"第 {i} 行: 手机号 '{phone_str}' 与第 {seen_phones[phone_str]} 行重复")
                continue
            seen_phones[phone_str] = i

            # 与数据库比对
            existing = session.exec(select(User).where(User.phone == phone_str)).first()
            if existing:
                errors.append(f"第 {i} 行: 手机号 '{phone_str}' 已被系统中的用户 '{existing.name}' 占用")
                continue

            user = User(
                name=str(name).strip(),
                phone=phone_str,
                email=str(email).strip() if email else None,
                district=str(district).strip() if district else None,
                department=str(dept).strip() if dept else None,
                is_active=True,
                password=str(password).strip()
            )
            users_to_add.append(user)

        if errors:
            raise HTTPException(status_code=400, detail={
                "message": f"导入失败，发现 {len(errors)} 个数据错误或冲突",
                "errors": errors
            })

        session.add_all(users_to_add)
        session.commit()

        return {"count": len(users_to_add), "message": f"新增成功了 {len(users_to_add)} 人"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")

@router.post("/", response_model=UserRead)
def create_user(user: User, session: Session = Depends(get_session)):
    """
    Create New User
    """
    # Check if phone exists (optional)
    if user.phone:
        existing = session.exec(select(User).where(User.phone == user.phone)).first()
        if existing:
            raise HTTPException(status_code=400, detail="Phone already registered")

    # 如果没有传密码，设置默认密码
    if not user.password:
        user.password = "password123"

    session.add(user)
    session.commit()
    session.refresh(user)
    return user

@router.get("/")
def read_users(
    session: Session = Depends(get_session),
    page: int = 1,
    page_size: int = 10,
    q: Optional[str] = None,
    is_active: Optional[bool] = None,
    districts: Optional[str] = Query(None, description="Comma-separated list of districts"),
    sort_by: Optional[str] = Query(None),
    sort_order: Optional[str] = Query(None)
):
    """
    Get Users List with Pagination and Filtering
    """
    
    # Efficient Count
    # We clone the query conditions for counting
    count_query = select(func.count()).select_from(User)
    if q:
        count_query = count_query.where((User.phone.contains(q)) | (User.name.contains(q)))
    if is_active is not None:
        count_query = count_query.where(User.is_active == is_active)
    if districts:
        dist_list = districts.split(",")
        count_query = count_query.where(User.district.in_(dist_list))
        
    total = session.exec(count_query).one()
    
    # Build Result Query
    query = select(User)
    if q:
        query = query.where((User.phone.contains(q)) | (User.name.contains(q)))
    if is_active is not None:
        query = query.where(User.is_active == is_active)
    if districts:
        dist_list = districts.split(",")
        query = query.where(User.district.in_(dist_list))

    # Sorting
    order_col = User.id.desc() # Default
    if sort_by == "district":
        order_col = User.district.asc() if sort_order == "ascending" else User.district.desc()
    elif sort_by == "department":
        order_col = User.department.asc() if sort_order == "ascending" else User.department.desc()
        
    query = query.offset((page - 1) * page_size).limit(page_size).order_by(order_col)
    items = session.exec(query).all()
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size
    }

@router.get("/{user_id}", response_model=UserRead)
def read_user(user_id: int, session: Session = Depends(get_session)):
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@router.put("/{user_id}", response_model=UserRead)
def update_user(user_id: int, user_data: User, session: Session = Depends(get_session)):
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Update fields
    user_data_dict = user_data.dict(exclude_unset=True)
    for key, value in user_data_dict.items():
        if key != 'id':  # Prevent ID change
            setattr(user, key, value)

    session.add(user)
    session.commit()
    session.refresh(user)
    return user

@router.delete("/{user_id}")
def delete_user(user_id: int, session: Session = Depends(get_session)):
    """
    Delete User by ID
    """
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    session.delete(user)
    session.commit()
    return {"ok": True}

@router.post("/change_password")
def change_password(req: ChangePasswordRequest, session: Session = Depends(get_session)):
    """
    修改用户密码
    """
    user = session.get(User, req.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # 验证旧密码（兼容明文旧密码和 bcrypt 哈希）
    current_pwd = user.password or "password123"  # Fallback to default if null
    if not verify_password(req.old_password, current_pwd):
        raise HTTPException(status_code=400, detail="旧密码错误")

    # 新密码使用明文存储
    user.password = req.new_password
    session.add(user)
    session.commit()

    return {"message": "密码修改成功"}
